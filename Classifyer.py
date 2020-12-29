#!/usr/bin/env pythonpy
from PIL import Image
from plantcv import plantcv as pcv
import cv2
import numpy as np
import os, shutil, glob, os.path
import pandas as pd
import matplotlib.pyplot as plt
from openpyxl import load_workbook
import glob
from imutils import paths
import argparse
import scipy.stats
import sys
from sklearn.cluster import KMeans
from sklearn.preprocessing import normalize
import scipy.cluster.hierarchy as shc
from sklearn.cluster import AgglomerativeClustering
import statistics
import concurrent.futures
import easygui

total_size_array = []
total_size_avg_array = []
total_size_std_array = []
total_color_array = []
total_color_avg_array = []
total_color_std_array = []
temp_array = []
temp_color = []
base_arr = []
platename_arr = []
Q1_size = []
Q2_size = []
Q3_size = []
Q4_size = []
Z1_size = []
Z2_size = []
Z3_size = []
Z4_size = []
Z_avg = []
above_size_ther = []
mod_size = []
temp = []
Q1_color = []
Q2_color = []
Q3_color = []
Q4_color = []
Z1_color = []
Z2_color = []
Z3_color = []
Z4_color = []
Z_avg_color = []
above_size_ther_color = []
mod_color = []
temp_color = []

image_counter = 0




def initexecl_liz(): #U platenumber - plate row A-H col number 1-12
    df = pd.DataFrame(
        {'Image processed': (), 'Cluster': (), 'Q1_size': (), 'Q2_size': (), 'Q3_size': (),
         'Q4_size': (),'Avg_size': (), 'Size_stdev': (), 'Q1_Zscore': (),'Q2_Zscore': (), 'Q3_Zscore': (),
         'Q4_Zscore': (), 'Zscore_avg': (), 'above_sizethres': (), 'modifier': (), 'temp': (),'Hit': (),
         'Q1_colorfullness': (), 'Q2_colorfullness': (), 'Q3_colorfullness': (), 'Q4_colorfullness': (),
         'Avg_color': (),'Color_stdev':(), 'Q1_color_Zscore' : (), 'Q2_color_Zscore': (), 'Q3_color_Zscore': (),
         'Q4_color_Zscore': (),'color_Zscore_avg': (), 'above color_thres': (), 'color_modifier': (), 'color_temp': (),
         'colorhit': (),'POSITIVE': ()})
    writer = pd.ExcelWriter("A_test.xlsx", engine='openpyxl')
    df.to_excel(writer,index=False, header=True, startcol=0)
    writer.save()
    # https://medium.com/better-programming/using-python-pandas-with-excel-d5082102ca27

def initexecl_chris(): #U platenumber - plate row A-H col number 1-12
    df = pd.DataFrame(
        {'Image processed': (), 'Cluster': (), 'Q1_size': (), 'Q2_size': (), 'Q3_size': (),
         'Q4_size': (),'Avg_size': (), 'Size_stdev': (), 'Q1_Zscore': (),'Q2_Zscore': (), 'Q3_Zscore': (),
         'Q4_Zscore': (), 'Zscore_avg': (), '# of threshold': (), 'modifier': (), 'temp': (),'POSITIVE': ()})
    writer = pd.ExcelWriter("A_test.xlsx", engine='openpyxl')
    df.to_excel(writer,index=False, header=True, startcol=0)
    writer.save()

print("Size and color processing 1 Size processing 2")
mode = easygui.indexbox(msg="What do you want to process\nNote can not process a single image folder must have two or more",
                        title="Yeast Classifier",
                 choices=("Size and color", "Size"))
if int(mode) == 0:
    initexecl_liz()
if int(mode) == 1:
    initexecl_chris()
folder = easygui.diropenbox()
imagePath = sorted(list(paths.list_images(folder)))
for i in range(len(imagePath)):
    img = Image.open(imagePath[i])
    #img.show()
    base = os.path.basename(imagePath[i])
    print("PROCESSING IMAGE %s..." %base)

    def initcrop(img): #change line 70 for threshold adj
        left = 1875  # was 2050
        top = 730  # was 870
        right = 5680
        bottom = 3260  # was 3280
        dire = os.getcwd()
        path = dire + '/Classifyer_dump'
        try:
            os.makedirs(path)
        except OSError:
            pass
        img_crop = img.crop((left, top, right, bottom))
        # img_crop.show()
        img_crop.save(os.path.join(path, 'Cropped_full_yeast.png'))
        circle_me = cv2.imread(os.path.join(path, "Cropped_full_yeast.png"))
        cropped_img = cv2.imread(
            os.path.join(path, "Cropped_full_yeast.png"))  # changed from Yeast_Cluster.%d.png  %counter
        blue_image = pcv.rgb2gray_lab(cropped_img, 'b')  # can do l a or b
        Gaussian_blue = cv2.adaptiveThreshold(blue_image, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C, cv2.THRESH_BINARY, 241,
                                              -1)  # For liz's pictures 241
        cv2.imwrite(os.path.join(path, "blue_test.png"), Gaussian_blue)
        blur_image = pcv.median_blur(Gaussian_blue, 10)
        heavy_fill_blue = pcv.fill(blur_image, 400)  # value 400
        cv2.imwrite(os.path.join(path, "Cropped_Threshold.png"), heavy_fill_blue)


    def cluster_maker():
        counter3 = 0
        counter1 = 0
        dire = os.getcwd()
        path = dire + '/Classifyer_dump'
        path1 = dire + '/Yeast_cluster_inv'
        path2 = dire + '/Yeast_cluster'
        path3 = dire + '/Cells'
        try:
            os.makedirs(path)
        except OSError:
            pass
        try:
            os.makedirs(path1)
        except OSError:
            pass
        try:
            os.makedirs(path2)
        except OSError:
            pass
        try:
            os.makedirs(path3)
        except OSError:
            pass
        counter = 0
        im = Image.open(os.path.join(path, "Cropped_Threshold.png"))  # was "Cropped_full_yeast.png"
        sizeX, sizeY = im.size
        im_sizeX = round(sizeX / 12)
        im_sizeY = round(sizeY / 8)
        for h in range(0, im.height, im_sizeY):
            nim = im.crop((0, h, im.width - 1, min(im.height, h + im_sizeY) - 1))
            nim.save(os.path.join(path, "Yeast_Row." + str(counter) + ".png"))
            counter += 1
        anotherCounter = 0
        for i in range(0, 8):
            columnImage = (os.path.join(path, "Yeast_Row.%d.png" % anotherCounter))
            Each_Image = Image.open(columnImage)
            sizeX2, sizeY2 = Each_Image.size
            Each_Image_sizeX = round(sizeX2 / 12)
            Each_Image_sizeY = round(sizeY2 / 8)
            anotherCounter += 1
            widthCounter1 = 0
            widthCounter2 = Each_Image_sizeX
            for w in range(0, 12):
                Wim = Each_Image.crop(
                    (widthCounter1, w, widthCounter2, min(Each_Image.height, w + Each_Image_sizeX) - 1))
                Wim.save(os.path.join(path1, "Yeast_Cluster." + str(counter1) + ".png"))
                counter1 += 1
                widthCounter1 = widthCounter1 + Each_Image_sizeX
                widthCounter2 = widthCounter2 + Each_Image_sizeX
                print(counter1)
        im = Image.open(os.path.join(path, "Cropped_full_yeast.png"))  # was "Cropped_full_yeast.png"
        counter = 0
        anotherCounter = 0
        counter1 = 0
        sizeX, sizeY = im.size
        im_sizeX = round(sizeX / 12)
        im_sizeY = round(sizeY / 8)
        for h in range(0, im.height, im_sizeY):
            nim = im.crop((0, h, im.width - 1, min(im.height, h + im_sizeY) - 1))
            nim.save(os.path.join(path, "Yeast_Row." + str(counter) + ".png"))
            counter += 1
        anotherCounter = 0
        for i in range(0, 8):
            columnImage = (os.path.join(path, "Yeast_Row.%d.png" % anotherCounter))
            Each_Image = Image.open(columnImage)
            sizeX2, sizeY2 = Each_Image.size
            Each_Image_sizeX = round(sizeX2 / 12)
            Each_Image_sizeY = round(sizeY2 / 8)
            anotherCounter += 1
            widthCounter1 = 0
            widthCounter2 = Each_Image_sizeX
            for w in range(0, 12):
                Wim = Each_Image.crop(
                    (widthCounter1, w, widthCounter2, min(Each_Image.height, w + Each_Image_sizeX) - 1))
                Wim.save(os.path.join(path2, "Yeast_Cluster." + str(counter1) + ".png"))
                counter1 += 1
                widthCounter1 = widthCounter1 + Each_Image_sizeX
                widthCounter2 = widthCounter2 + Each_Image_sizeX
        row_counter_for_save = 0
        row_counter_for_open = 0
        for i in range(0, 96):
            im = Image.open(os.path.join(path2, "Yeast_Cluster.%d.png" % i))
            sizeX, sizeY = im.size
            im_sizeX = round(sizeX / 2)
            im_sizeY = round(sizeY / 2)
            for h in range(0, im.height, im_sizeY):
                nim = im.crop((0, h, im.width - 1, min(im.height, h + im_sizeY) - 1))
                nim.save(os.path.join(path, "ROW_SMALL." + str(row_counter_for_save) + ".png"))
                row_counter_for_save += 1
                if (h >= im_sizeY):
                    break
            for i in range(0, 2):
                rowImage = (os.path.join(path, "ROW_SMALL.%d.png" % row_counter_for_open))
                Each_Image = Image.open(rowImage)
                sizeX2, sizeY2 = Each_Image.size
                Each_Image_sizeX = round(sizeX2 / 2)
                Each_Image_sizeY = round(sizeY2 / 2)
                row_counter_for_open += 1
                widthCounter1 = 0
                widthCounter2 = Each_Image_sizeX
                for w in range(0, 2):
                    Wim = Each_Image.crop(
                        (widthCounter1, w, widthCounter2, min(Each_Image.height, w + Each_Image_sizeX) - 1))
                    Wim.save(os.path.join(path3, "SMALL_CELL." + str(counter3) + ".png"))
                    counter3 += 1
                    widthCounter1 = widthCounter1 + Each_Image_sizeX
                    widthCounter2 = widthCounter2 + Each_Image_sizeX


    def connected_comps_for_liz(counter):
        dire = os.getcwd()
        path = dire + '/Yeast_cluster_inv'
        cropped_img = cv2.imread(os.path.join(path, 'Yeast_Cluster.%d.png' % counter),
                                 cv2.IMREAD_UNCHANGED)  # changed from Yeast_Cluster.%d.png  %counter
        circle_me = cv2.imread(os.path.join(path, "Yeast_Cluster.%d.png" % counter))

        connected_counter = 0

        connectivity = 8

        connectivity = 8  # either 4 or 8
        output = cv2.connectedComponentsWithStats(cropped_img, connectivity)  # Will determine the size of our clusters

        num_labels = output[0]
        labels = output[1]
        stats = output[2]  # for size
        centroids = output[3]  # for location

        #print("Currently on cell %d" % counter)
        cc_size_array = []
        #print(centroids)
        for i in range(0, (len(stats)), 1):
            if (centroids[i][0] >= 40 and centroids[i][0] <= 110 and centroids[i][1] >= 40 and centroids[i][1] <= 110):
                print("%d is in 1" % i)
                cc_size_array.append(stats[i, cv2.CC_STAT_AREA])
        for i in range(0, (len(stats)), 1):
            if (centroids[i][0] >= 200 and centroids[i][0] <= 270 and centroids[i][1] >= 40 and centroids[i][1] <= 110):
                cc_size_array.append(stats[i, cv2.CC_STAT_AREA])
                print("%d is in 2" % i)
        for i in range(0, (len(stats)), 1):
            if (centroids[i][0] >= 40 and centroids[i][0] <= 110 and centroids[i][1] >= 200 and centroids[i][1] <= 270):
                cc_size_array.append(stats[i, cv2.CC_STAT_AREA])
                print("%d is in 3" % i)
        for i in range(0, (len(stats)), 1):
            if (centroids[i][0] >= 200 and centroids[i][0] <= 270 and centroids[i][1] >= 200 and centroids[i][
                1] <= 270):
                cc_size_array.append(stats[i, cv2.CC_STAT_AREA])
                print("%d is in 4" % i)

        if (len(stats) < 4):
            print("too few decteted on %d" % counter)
            print((len(stats)))
            for i in range((len(stats)), 5, 1):
                cc_size_array.append(0)

        if (len(cc_size_array) >= 5):
            print("problem on cell %d" % counter)
            exit(-1)

        # total_size_array = total_size_array + cc_size_array
        #print("size data")
        #print(cc_size_array)
        avg_size = np.average(cc_size_array)
        #print(avg_size)
        std = np.std(np.array(cc_size_array))
        #print(std)
        Zscore_array = abs(scipy.stats.zscore(cc_size_array))
        #print(Zscore_array)
        Z_avg = np.average(Zscore_array)
        #print(Z_avg)
        mod = 1.5  # if avg zscore is less than .5 or 40% that is rewarded
        if Z_avg >= .8: # completely random number lol
            mod = 1
        above_size_ther = 0
        for i in range(0, len(cc_size_array)):
            if cc_size_array[i] <= 3500:  # for liz she wants small?
                above_size_ther += 1

        temp = ((10*above_size_ther)-(mod * Z_avg))  # simply alg to tell is positive gets normalized later
        #print(temp)


        #print("end of size data")



        return cc_size_array, avg_size, std, Zscore_array, Z_avg, above_size_ther, mod, temp


    def connected_comps_for_Chris(counter):
        dire = os.getcwd()
        path = dire + '/Yeast_cluster_inv'
        cropped_img = cv2.imread(os.path.join(path, 'Yeast_Cluster.%d.png' % counter),
                                 cv2.IMREAD_UNCHANGED)  # changed from Yeast_Cluster.%d.png  %counter
        circle_me = cv2.imread(os.path.join(path, "Yeast_Cluster.%d.png" % counter))

        connected_counter = 0

        connectivity = 8

        connectivity = 8  # either 4 or 8
        output = cv2.connectedComponentsWithStats(cropped_img, connectivity)  # Will determine the size of our clusters

        num_labels = output[0]
        labels = output[1]
        stats = output[2]  # for size
        centroids = output[3]  # for location

        #print("Currently on cell %d" % counter)
        cc_size_array = []
        #print(centroids)
        for i in range(0, (len(stats)), 1):
            if (centroids[i][0] >= 40 and centroids[i][0] <= 110 and centroids[i][1] >= 40 and centroids[i][1] <= 110):
                print("%d is in 1" % i)
                cc_size_array.append(stats[i, cv2.CC_STAT_AREA])
        for i in range(0, (len(stats)), 1):
            if (centroids[i][0] >= 200 and centroids[i][0] <= 270 and centroids[i][1] >= 40 and centroids[i][1] <= 110):
                cc_size_array.append(stats[i, cv2.CC_STAT_AREA])
                print("%d is in 2" % i)
        for i in range(0, (len(stats)), 1):
            if (centroids[i][0] >= 40 and centroids[i][0] <= 110 and centroids[i][1] >= 200 and centroids[i][1] <= 270):
                cc_size_array.append(stats[i, cv2.CC_STAT_AREA])
                print("%d is in 3" % i)
        for i in range(0, (len(stats)), 1):
            if (centroids[i][0] >= 200 and centroids[i][0] <= 270 and centroids[i][1] >= 200 and centroids[i][
                1] <= 270):
                cc_size_array.append(stats[i, cv2.CC_STAT_AREA])
                print("%d is in 4" % i)

        if (len(stats) < 4):
            print("too few decteted on %d" % counter)
            print((len(stats)))
            for i in range((len(stats)), 5, 1):
                cc_size_array.append(0)

        if (len(cc_size_array) >= 5):
            print("problem on cell %d" % counter)
            exit(-1)

        # total_size_array = total_size_array + cc_size_array
        #print("size data")
        #print(cc_size_array)
        avg_size = np.average(cc_size_array)
        #print(avg_size)
        std = np.std(np.array(cc_size_array))
        #print(std)
        Zscore_array = abs(scipy.stats.zscore(cc_size_array))
        #print(Zscore_array)
        Z_avg = np.average(Zscore_array)
        #print(Z_avg)
        mod = 1.5  # if avg zscore is less than .5 or 40% that is rewarded
        if Z_avg >= .8:  # completely random number lol
            mod = 1
        above_size_ther = 0
        for i in range(0, len(cc_size_array)):
            if cc_size_array[i] >= 3500:  # for Chris likes them big
                above_size_ther += 1

        temp = ((10 * above_size_ther) - (mod * Z_avg))  # simply alg to tell is positive gets normalized later
        #print(temp)

        #print("end of size data")

        return cc_size_array, avg_size, std, Zscore_array, Z_avg, above_size_ther, mod, temp


    def image_colorfulness(image):
        # split the image into its respective RGB components
        (B, G, R) = cv2.split(image.astype("float"))
        # compute rg = R - G
        rg = np.absolute(R - G)
        # compute yb = 0.5 * (R + G) - B
        yb = np.absolute(0.5 * (R + G) - B)
        # compute the mean and standard deviation of both `rg` and `yb`
        (rbMean, rbStd) = (np.mean(rg), np.std(rg))
        (ybMean, ybStd) = (np.mean(yb), np.std(yb))
        # combine the mean and standard deviations
        stdRoot = np.sqrt((rbStd ** 2) + (ybStd ** 2))
        meanRoot = np.sqrt((rbMean ** 2) + (ybMean ** 2))
        # derive the "colorfulness" metric and return it
        return stdRoot + (0.3 * meanRoot)


    def colorful_writer(color_counter):
        dire = os.getcwd()
        path = dire + '/Cells'
        # https://www.pyimagesearch.com/2017/06/05/computing-image-colorfulness-with-opencv-and-python/
        color_array = []
        for i in range(0, 4):
            #print("THIS IS COLOR COUNTER")
            #print(color_counter)
            image = cv2.imread(os.path.join(path, "SMALL_CELL.%d.png" % color_counter))
            C = image_colorfulness(image)
            # display the colorfulness score on the image
            color_array.append(C)
           # cv2.putText(image, "{:.2f}".format(C), (40, 40), cv2.FONT_HERSHEY_SIMPLEX, 1, (0, 255, 0), 3)
            #cv2.imwrite(os.path.join(path, "SMALL_CELL.%d.png" % color_counter), image)
            color_counter = color_counter + 1

        # total_color_array = total_color_array + color_array
        avg_color = np.average(color_array)
        std_color = np.std(np.array(color_array))

        if (len(color_array) < 4):
            print((len(color_array)))
            for i in range((len(color_array)), 5, 1):
                color_array.append(0)

        Zscore_array = abs(scipy.stats.zscore(color_array))
        #print(Zscore_array)
        Z_avg = np.average(Zscore_array)
        #print(Z_avg)
        mod = 1.5  # if avg zscore is less than .5 or 40% that is rewarded
        if Z_avg >= .8:  # completely random number lol
            mod = 1
        above_size_ther = 0
        for i in range(0, len(color_array)):
            if color_array[i] <= 23: #pretty white ones???
                above_size_ther += 1
        temp = ((10*above_size_ther) - (mod * Z_avg))


        #print(color_array)
        return color_array, avg_color, std_color, color_counter, Zscore_array, Z_avg, above_size_ther, mod, temp


    def size_hit(temp_array):
        X = temp_array
        X_max = max(X)
        normalize_array = []
        for i in range(0, (image_counter * 96)):
            x = (X[i] / X_max) * 100
            if x > 70:
                x = 1
            else:
                x = 0
            normalize_array.append(x)
        return normalize_array


    def color_hit(temp_color):
        normalize_array_color = []
        X = temp_color
        X_max = max(X)
        for i in range(0, (image_counter * 96)):
            x = (X[i] / X_max) * 100
            if x > 70:
                x = 1
            else:
                x = 0
            normalize_array_color.append(x)
        return normalize_array_color


    def pos_hit(normalize_array, normalize_array_color):
        pos_array = []
        for i in range(0, (image_counter * 96)):
            if normalize_array[i] == 1 and normalize_array_color[i] == 1:
                pos_array.append(1)
            else:
                pos_array.append(0)
        return pos_array


    def excel_writer_liz(base_arr, platename_arr, Q1_size, Q2_size, Q3_size, Q4_size,
                         total_size_avg_array, total_size_std_array, Z1_size, Z2_size, Z3_size, Z4_size, Z_avg,
                         above_size_ther, mod_size, temp_array, Q1_color, Q2_color, Q3_color, Q4_color,
                         total_color_avg_array, total_color_std_array, Z1_color, Z2_color, Z3_color,
                         Z4_color, above_size_ther_color, mod_color, temp_color, size_pos, color_pos, pos_size):
        new_df = pd.DataFrame(
            {'Image processed':(base_arr),'Cluster' :(platename_arr),
             'Q1_size': (Q1_size), 'Q2_size': (Q2_size),'Q3_size': (Q3_size),
             'Q4_size': (Q4_size),'Avg_size': (total_size_avg_array), 'Size_stdev': (total_size_std_array),
             'Q1_Zscore': (Z1_size), 'Q2_Zscore': (Z2_size), 'Q3_Zscore': (Z3_size),
             'Q4_Zscore': (Z4_size),'Avg_Zscore': (Z_avg), 'above_sizethres': (above_size_ther),
             'modifier': (mod_size), 'temp': (temp_array), 'hit': (size_pos), 'Q1_color': (Q1_color), 'Q2_color': (Q2_color)
                , 'Q3_color': (Q3_color), 'Q4_color': (Q4_color), 'Avg_color': (total_color_avg_array),
             'Color_stdev': (total_color_std_array), 'Q1_color_Zscore' : (Z1_color), 'Q2_color_Zscore': (Z2_color),
             'Q3_color_Zscore': (Z3_color),'Q4_color_Zscore': (Z4_color),
             'color_Zscore_avg': (Z_avg_color), 'above color_thres': (above_size_ther_color), 'modifier_color': (mod_color),
             'color_temp': (temp_color), 'colorhit': (color_pos), 'POSITIVE': (pos_size)})
        #new_df = pd.DataFrame.from_dict(new_df1, orient='index') # this is weird fuck indent for some works but columns does not
        #new_df.transpose()
        writer = pd.ExcelWriter('A_test.xlsx', engine='openpyxl')
        writer.book = load_workbook('A_test.xlsx')
        writer.sheets = dict((ws.title, ws) for ws in writer.book.worksheets)
        reader = pd.read_excel(r'A_test.xlsx')
        new_df.to_excel(writer, index=False, header=False, startcol=0, startrow=len(reader) + 1)
        writer.close()


    def excel_writer_chris(base_arr, platename_arr, Q1_size, Q2_size, Q3_size, Q4_size,
                         total_size_avg_array, total_size_std_array, Z1_size, Z2_size, Z3_size, Z4_size, Z_avg,
                         above_size_ther, mod_size, temp_array, pos_size):
        new_df = pd.DataFrame(
            {'Image processed':(base_arr),'Cluster' :(platename_arr),
             'Q1_size': (Q1_size), 'Q2_size': (Q2_size),'Q3_size': (Q3_size),
             'Q4_size': (Q4_size),'Avg_size': (total_size_avg_array), 'Size_stdev': (total_size_std_array),
             'Q1_Zscore': (Z1_size), 'Q2_Zscore': (Z2_size), 'Q3_Zscore': (Z3_size),
             'Q4_Zscore': (Z4_size),'Avg_Zscore': (Z_avg), '# above threshold': (above_size_ther),
             'modifier': (mod_size), 'temp': (temp_array), 'hit': (pos_size)})
        writer = pd.ExcelWriter('A_test.xlsx', engine='openpyxl')
        writer.book = load_workbook('A_test.xlsx')
        writer.sheets = dict((ws.title, ws) for ws in writer.book.worksheets)
        reader = pd.read_excel(r'A_test.xlsx')
        new_df.to_excel(writer, index=False, header=False, startcol=0, startrow=len(reader) + 1)
        writer.close()


    toomanycounter = 1
    anothercounter = 1
    color_counter = 0
    plate_size = []
    initcrop(img)
    cluster_maker()
    if int(mode) == 0:
        print("liz")
        image_counter = image_counter + 1
        for c in range(0, 96):
            base_arr.append(base)
            char = chr(toomanycounter + 64)
            plate_name = ("U%d-%c%d" % (image_counter, char, anothercounter))
            platename_arr.append(plate_name)
            returned_size = connected_comps_for_liz(c) #inputs is counter for which cluster to process and output is an array with size, avg size, and std
            #print(returned_size)
            returned_color = colorful_writer(color_counter) #input is color_counter so knows which cell to process output is an array with colorfulness, avg color, and std
            anothercounter = anothercounter + 1
            if anothercounter > 12:
                anothercounter = 1
                toomanycounter = toomanycounter + 1
            color_counter = returned_color[3]
            #cc_size_array, avg_size, std, Zscore_array, Z_avg, above_size_ther, mod, temp
            cc = []
            cc = returned_size[0]
            #print(cc)
            Q1_size.append(cc[0])
            Q2_size.append(cc[1])
            Q3_size.append(cc[2])
            Q4_size.append(cc[3])
            total_size_avg_array.append(returned_size[1])
            total_size_std_array.append(returned_size[2])
            Zscore_array = (returned_size[3])
            Z1_size.append(Zscore_array[0])
            Z2_size.append(Zscore_array[1])
            Z3_size.append(Zscore_array[2])
            Z4_size.append(Zscore_array[3])
            Z_avg.append(returned_size[4])
            above_size_ther.append(returned_size[5])
            mod_size.append(returned_size[6])
            temp_array.append(returned_size[7])
            plate_size.extend(returned_size[0])
            color_array = []
            color_array = (returned_color[0])
            Q1_color.append(color_array[0])
            Q2_color.append(color_array[1])
            Q3_color.append(color_array[2])
            Q4_color.append(color_array[3])
            total_color_avg_array.append(returned_color[1])
            total_color_std_array.append(returned_color[2])
            Z_color = []
            Z_color = returned_color[4]
            Z1_color.append(Z_color[0])
            Z2_color.append(Z_color[1])
            Z3_color.append(Z_color[2])
            Z4_color.append(Z_color[3])
            Z_avg_color.append(returned_color[5])
            above_size_ther_color.append(returned_color[6])
            mod_color.append(returned_color[7])
            temp_color.append(returned_color[8])

        if(image_counter > 2): #changed for 2
            break
        #print(platename_arr)
    if int(mode) == 1:
        print("Chris")
        image_counter = image_counter + 1
        for c in range(0, 96):
            base_arr.append(base)
            char = chr(toomanycounter + 64)
            plate_name = ("U%d-%c%d" % (image_counter, char, anothercounter))
            platename_arr.append(plate_name)
            returned_size = connected_comps_for_Chris(c)  # inputs is counter for which cluster to process and output is an array with size, avg size, and std
            # print(returned_size)
            returned_color = colorful_writer(color_counter)  # input is color_counter so knows which cell to process output is an array with colorfulness, avg color, and std
            anothercounter = anothercounter + 1
            if anothercounter > 12:
                anothercounter = 1
                toomanycounter = toomanycounter + 1
            color_counter = returned_color[3]
            cc = []
            cc = returned_size[0]
            # print(cc)
            Q1_size.append(cc[0])
            Q2_size.append(cc[1])
            Q3_size.append(cc[2])
            Q4_size.append(cc[3])
            total_size_avg_array.append(returned_size[1])
            total_size_std_array.append(returned_size[2])
            Zscore_array = (returned_size[3])
            Z1_size.append(Zscore_array[0])
            Z2_size.append(Zscore_array[1])
            Z3_size.append(Zscore_array[2])
            Z4_size.append(Zscore_array[3])
            Z_avg.append(returned_size[4])
            above_size_ther.append(returned_size[5])
            mod_size.append(returned_size[6])
            temp_array.append(returned_size[7])
            plate_size.extend(returned_size[0])
            temp_color.append(returned_color[8])
        if (image_counter > 2):
            break

if int(mode) == 0:
    size_pos = size_hit(temp_array)
    color_pos = color_hit(temp_color)
    pos_size = pos_hit(size_pos, color_pos)
    excel_writer_liz(base_arr, platename_arr, Q1_size, Q2_size, Q3_size, Q4_size,
                     total_size_avg_array, total_size_std_array, Z1_size, Z2_size, Z3_size, Z4_size, Z_avg,
                     above_size_ther, mod_size, temp_array, Q1_color, Q2_color, Q3_color, Q4_color,
                     total_color_avg_array, total_color_std_array, Z1_color, Z2_color, Z3_color,
                     Z4_color, above_size_ther_color, mod_color, temp_color, size_pos, color_pos, pos_size)
if int(mode) == 1:
    pos_size = size_hit(temp_array)
    excel_writer_chris(base_arr, platename_arr, Q1_size, Q2_size, Q3_size, Q4_size,
                       total_size_avg_array, total_size_std_array, Z1_size, Z2_size, Z3_size, Z4_size, Z_avg,
                       above_size_ther, mod_size, temp_array, pos_size)

easygui.msgbox(msg="DONE!!!", title="Yeast Classifier")

"""
#beep = lambda x: os.system("echo -n '\a';sleep 0.2;" * x)
#beep(5)

X = temp_array
X_max = max(X)
(X_max)
normalize_array = []
normalize_array_color = []
pos_array = []
for i in range(0, (image_counter*96)):
    x = (X[i]/X_max) * 100
    if x > 70:
        x = 1
    else:
        x = 0
    normalize_array.append(x)

    X = temp_color
    X_max = max(X)
    #print(X_max)
    x = (X[i]/X_max) * 100
    if x > 70:
        x = 1
    else:
        x = 0
    normalize_array_color.append(x)

    if normalize_array[i] == 1 and normalize_array_color[i] == 1:
        pos_array.append(1)
    else:
        pos_array.append(0)

    new_df = pd.DataFrame(
                {'Hit': (normalize_array[i])}, index=[0])
    new_df2 = pd.DataFrame(
                {'colorhit': (normalize_array_color[i])}, index=[0])
    new_df3 = pd.DataFrame(
                {'POSITIVE': (pos_array[i])}, index=[0])
    writer = pd.ExcelWriter('A_test.xlsx', engine='openpyxl')
    writer.book = load_workbook('A_test.xlsx')
    writer.sheets = dict((ws.title, ws) for ws in writer.book.worksheets)
    reader = pd.read_excel(r'A_test.xlsx')
    new_df.to_excel(writer, sheet_name='Sheet1', index=False, header=False, startcol=16, startrow=(i + 1))
    new_df2.to_excel(writer, index=False, header=False, startcol=30, startrow=(i + 1))
    new_df3.to_excel(writer, index=False, header=False, startcol=31, startrow=(i + 1))
    writer.close()
#print(pos_array)

"""

"""
#kmeans stuff
imdir = '/Users/gregglickert/Documents/GitHub/YeastClassification/Cluster'
targetdir = "/Users/gregglickert/Documents/GitHub/YeastClassification/cluster_test"
filelist = glob.glob(os.path.join(imdir, '*.png'))
filelist.sort()
try:
    os.makedirs(targetdir)
except OSError:
    pass
# Copy with cluster name
print("\n")
for i, m in enumerate(kmeans.labels_): #changed from kmeans to dbscan
    print("    Copy: %s / %s" %(i, len(kmeans.labels_)), end="\r") #same here
    shutil.move(filelist[i], '/Users/gregglickert/Documents/GitHub/YeastClassification/cluster_test')
"""
#His clustering
"""
plt.figure(figsize=(10, 7))
plt.title("Dendrograms")
#dend = shc.dendrogram(shc.linkage(X, method='ward'))
model = AgglomerativeClustering(n_clusters=3, affinity='euclidean', linkage='ward')
model.fit(X)
labels = model.labels_
plt.scatter(X[labels==0, 0], X[labels==0, 1], s=50, marker='o', color='red')
plt.scatter(X[labels==1, 0], X[labels==1, 1], s=50, marker='o', color='blue')
plt.scatter(X[labels==2, 0], X[labels==2, 1], s=50, marker='o', color='green')
plt.scatter(X[labels==3, 0], X[labels==3, 1], s=50, marker='o', color='purple')
plt.scatter(X[labels==4, 0], X[labels==4, 1], s=50, marker='o', color='orange')
plt.show()
"""






"""
x = data[["Avg_size","Avg_color"]]
plt.scatter(x["Avg_size"],x["Avg_color"])
plt.xlabel('Avg_size')
plt.ylabel('Avg_color')
plt.show()
"""
#https://www.analyticsvidhya.com/blog/2019/08/comprehensive-guide-k-means-clustering/

"""
for i in range(0,96):
    cluster_array[i] = ([cluster_size_array[i] + cluster_color_std_array[i] + cluster_color_array[i] + cluster_size_std_array[i]])
print("clustered is")
print(cluster_array)
print(len(cluster_array))
imdir = '/Users/gregglickert/PycharmProjects/cc_test/CLUSTERD'
targetdir = "/Users/gregglickert/PycharmProjects/cc_test/testing folder"
number_clusters = 5
# Loop over files and get features
filelist = glob.glob(os.path.join(imdir, '*.png'))
filelist.sort()
kmeans = KMeans(n_clusters=number_clusters, random_state=0).fit(np.array(total_size_array))
try:
    os.makedirs(targetdir)
except OSError:
    pass
# Copy with cluster name
print("\n")
for i, m in enumerate(kmeans.labels_): #changed from kmeans to dbscan
    print("    Copy: %s / %s" %(i, len(kmeans.labels_)), end="\r") #same here
    shutil.move(filelist[i], '/Users/gregglickert/PycharmProjects/cc_test/testing folder2')
print(len(total_size_array))
print(total_size_array)
print(len(total_color_array))
print(total_color_array)
plt.hist(total_size_array)
plt.show()
"""
